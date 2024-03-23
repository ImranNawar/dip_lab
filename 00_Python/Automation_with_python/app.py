# Automation with Python
"""
Created on Mon Mar 18 14:10:44 2024

@author: Imran Nawar
"""
import openpyxl as xl
# openpyxl is a Python library used for working with Excel files (.xlsx extension). It allows you to create, read, and modify Excel files programmatically using Python
from openpyxl .chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet['a1']
    cell = sheet.cell(1,1)

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
        
    values = Reference(sheet, 
              min_row=2, 
              max_row=sheet.max_row, 
              min_col=4,
              max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)
    
process_workbook('transactions.xlsx')

